import requests   #support for pulling contents of webpages
from bs4 import BeautifulSoup  #function for reading the page's XML returned by requests library
import re    #support for regular expressions
from time import sleep  #sleep function allows pausing of script, to avoid getting rate-limited by BGG
import unicodecsv as csv  #use the unicodecsv library instead. Straight clone of CSV functions, but with unicode handling
from random import randint  #generate random integers, used in randomizing wait time
import openpyxl #used in creating and writing to .xlsx filtes
from pathlib import Path #used in handling Path objects
import os #used in reading and creating directory for output
import sys #used to quit script

#########################################################
# Function to extract info from BGG and perform file I/O
#########################################################

def BGGextract():
    
    ####################################################
    # Configure directory and file in which to save info
    ####################################################

    #DEPRECATED - Set working directory by checking for BGGexports folder and creating it if does not exist
    # BGGpath = Path.home() / 'Documents' / 'BGGexports'
    # if BGGpath.exists() == False:
    #    os.makedirs(BGGpath)
    # os.chdir(BGGpath)
    filename = 'BGG_IDs.xlsx'

    #Check if file exists, then either open it or create a new blank workbook, and set active sheet. If created new wb, configure sheet name and headers
    if Path(filename).is_file():
        print ('Loading existing BGG ID# index for updating... [This may take some time, ~20s]')
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        row_counter = sheet.max_row + 1 #Initialize incremental value for last written row. 
        #This is necessary as some BGG IDs will be skipped, and there will not be a 1-for-1 relationship to past written ID#s to # of rows written
        print('# of rows previously written: ' + str(row_counter))

        #For existing files, set the first item by reading in BGG ID# of last written row. Set last item by calling function to extract highest possible value from GeekFeed RSS
        first_item = sheet.cell(row = sheet.max_row, column = 1).value + 1
        last_item = BGGmaxitem()
        #last_item = 20000 + sheet.cell(row = sheet.max_row, column = 1).value
        print('Script will extract IDs beginning with #' + str(first_item) + ' and ending with #' + str(last_item))
        print('\n') 
        
    else:
        choice = input('File does not exist. Would you like to create a new workbook? Warning: This script will take up to 24 hours to execute. [Y/N]: ')
        if choice.lower() == 'y':
            print('Creating new workbook...' + '\n')
            wb = openpyxl.Workbook()
            sheet = wb.active
            first_item = 1
            last_item = BGGmaxitem()

            #Configure new file's sheet title and header row.
            sheet.title = 'BGG Game IDs'
            sheet['A1'] = 'BGG ID'
            sheet['B1'] = 'Game Title'
            sheet['C1'] = 'Year Published'
            sheet['D1'] = 'Category'
            row_counter = 1 # Set rowcounter flag to begin at start of sheet. Acts as global variable to hold position of future row batches to be written.
        else:
            print('Returning to main menu...')
            sleep(2)
            return

    ######################################################
    # Perform exraction of info from BGG and write to file
    ######################################################

    base_url = 'https://www.boardgamegeek.com/xmlapi2/thing?id='
    
    #BGG XML API can handle 100 parameters at a time. Iterate through first_item/last_item range 100 values at a time.
    for x in range(int(first_item),int(last_item),100):  
        if x >= (last_item - 100):   #check to see if fewer than 100 values remain in range. If so, trim ID_range for last execution of loop
            ID_range = range(x,last_item + 1)
        else:
            ID_range = range(x,x+100)

        #configure URL by turning (x,x+100), a list of 100 values, into a sequence of strings, joined by commas. Then, append that to base_url
        URL_args = ','.join(list(map(str,ID_range))) 
        url = base_url + URL_args  

        #Use requests and BeautifulSoup to extract and read XML. Separaetly pull XML tags: (1) of <name> with type "primary", (2) of <item>
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'lxml')
        soup_items = soup.find_all('item')
        
        batch_rows = 0

        #iterate through the game name and ID# pairs, using attributes to extract game names and ID#
        for game in soup_items:    

            game_id_num = game.attrs['id']  #extracts the id='string' from any <item> tags
            category = game.attrs['type']
            
            # Further refine soup opject to find primary game name. Only extract the game name value if the soup.find succesfully returns a result
            soup_name = game.find('name', type = 'primary')

            if soup_name is not None:
                game_name = soup_name.attrs['value']

                # Not all 'item' tags in BGG have a yearpublished value, so assign 0 where it is missing
                soup_year = game.find('yearpublished')
                if (soup_year is not None) and (soup_year.attrs['value'] != ''):
                    year_published = soup_year.attrs['value']
                else:
                    year_published = 0
                
                # Text formatting for game name
                title = game_name.replace('&amp;', '&') # find and replace to correct HTML ampersand escaping
                title = title.strip('"') # stripping of leading and trailing double quotes, which appear inconsistently in BGG entries (no explanation)

                print(game_id_num + ' ' + title + ' [' + str(year_published) + ']') #output to terminal to monitor proress

                # Assign values to spreadsheet cells
                sheet.cell(row = batch_rows + row_counter, column = 1).value = int(game_id_num)
                if category != 'videogame':
                    sheet.cell(row = batch_rows + row_counter, column = 2).value = str(title)
                else:
                    sheet.cell(row = batch_rows + row_counter, column = 2).value = 'video game title removed'
                sheet.cell(row = batch_rows + row_counter, column = 3).value = int(year_published)
                sheet.cell(row = batch_rows + row_counter, column = 4).value = str(category)

                # Increment counter for number of rows written (out of the 100 ID#s iterated through in each batch)
                batch_rows += 1

        row_counter += batch_rows #Add this loop iteration's row counting value to the existing total row count variable
        
        print('\n' + 'Attempting to load next batch of BGG IDs. Will take 10-15 seconds...' '\n')        
        wb.save(str(filename))   #saves the file
        sleep(randint(10,15))  #sleep to prevent rate-limit or DOS


    print('BGG Extract has completed, and file has been saved')
    return
    

########################################################################
# Function to determine highest ID# available on BGG's GeekFeed RSS page
########################################################################

def BGGmaxitem():
    
    #set API URL for the GeekFeed page. Pull XML with requests.get and parse with BeautifulSoup
    url = 'https://boardgamegeek.com/recentadditions/rss?subdomain=&infilters%5B0%5D=thing&domain=boardgame'
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml')
    soup_names = soup.find_all('item', limit=None) #returns all <item> tags
    
    game_id_num = 0 #initialize variable that will be compared against potentially higher ID #s

    #iterate through all <item> tags, extracting the 6-digit BGG ID# via regex and setting value of game_id_num every time a higher ID# is found
    for game in soup_names:
        ID_regex = re.compile(r'(?<=guid>https://boardgamegeek.com/boardgame/)\d\d\d\d\d\d')  #define regex that strips out all but the six BGG ID# digits
        id_num = ID_regex.search(str(game))   #apply the regex
        if (type(id_num) == re.Match):  #Ensure valid # was returned. For non boardgame entries on BGG, the regex will fail and not create an re.Match object. Could also use 'is not None' here
            if (int(game_id_num) <= int(id_num.group())):
                game_id_num = int(id_num.group())
    
    return(game_id_num)


############################################
# Code to execute if script is run directly 
############################################

if __name__ == "__main__":
    BGGextract()